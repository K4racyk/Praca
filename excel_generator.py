import openpyxl
from openpyxl import Workbook
import os  # Dodanie importu os

def generate_excel(report, filename):
    # Sprawdź, czy plik już istnieje
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["LP", "Opis", "Kategoria", "Podkategoria", "Autor"])  # Nagłówki

    # Znalezienie pierwszego wolnego wiersza
    row = sheet.max_row + 1
    sheet.append([row-1, report['description'], report['category'], report['subcategory'], report['author']])

    wb.save(filename)

def load_user_data(filename, author):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Pomijamy nagłówki
        if row[4] == author or author == "manager":  # Sprawdzamy, czy to dane autora lub managera
            data.append(row)

    return data
